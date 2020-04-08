import utm
from openpyxl import load_workbook
from dbfread import DBF
import json


def parse_dbf(dbf_file_dir):
    coord_info = {}
    with DBF(dbf_file_dir) as table:
        for record in table:
            zone_n = 19  
            lon, lat = utm.to_latlon(int(record["coord_este"]), int(record["coord_nort"]), zone_n, northern=False, strict=False)
            coord = "{0},{1}".format(lon,lat)
            coord_info[record["nombre"].lower()] = coord

    return coord_info        

def parse_xlsm(xlsm_file_dir, coord_info, interval = 1):
    wb = load_workbook(filename= xlsm_file_dir, read_only=True)
    ws = wb['Caudales']
    table_index = ws[5]
    caudales_info = []
    for row in ws.rows:
        if row[0].value : 
            central_name = row[0].value.lower().replace("_", " ")
            if central_name in coord_info : 
                coord_entry_info = coord_info[central_name]
                week = 1
                for cell_index in list(range(2,50))[::interval]:
                    
                    new_entry = {"central":central_name,
                        "geo": coord_entry_info,
                        "caudal": row[cell_index].value,
                        "mes": table_index[cell_index].value,
                        "week": week,
                        "año": row[1].value
                     }
                    week += 1
                    if week%5 == 0:
                        week = 1
                    caudales_info.append(new_entry)
    return caudales_info       



if __name__ == "__main__":

    dbf_file_dir = '/home/francisco/Documents/Ipre verano 2020/hidroelectricas/centrales_hidroele_20181016234835.dbf'
    xlsm_file_dir = "/home/francisco/Documents/Ipre verano 2020/IPLP20191201/IPLP20191201.xlsm"
    interval = 1
    print("procesando dbf")
    coord_info = parse_dbf(dbf_file_dir)
    #print(coord_info)
    print("procesando xlsm")
    caudales_info = parse_xlsm(xlsm_file_dir, coord_info, interval)
    print(len(caudales_info))
    print(caudales_info[1])
    with open("PLP_with_coords_{0}.json".format(interval), 'w') as outputfile:
        json.dump(caudales_info, outputfile)